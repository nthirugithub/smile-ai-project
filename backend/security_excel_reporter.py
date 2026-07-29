import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MODULE_MAPPING = {
    "SEC_AUTH": "Module 1: Authentication & Password Policy Controls",
    "SEC_SESS": "Module 2: Session Management & JWT Token Boundaries",
    "SEC_AUTHZ": "Module 3: Authorization & Access Control Policies",
    "SEC_INP": "Module 4: Input Sanitization & Parameter Boundaries",
    "SEC_HDR": "Module 5: Security Headers & Transport Hardening",
    "SEC_PRIV": "Module 6: Data Privacy & Anonymization Controls",
    "SEC_ERR": "Module 7: Error Handling & Information Leakage Prevention",
    "SEC_LOG": "Module 8: Audit Logging & Security Event Integrity",
}

def extract_module_name(test_id):
    for prefix, name in MODULE_MAPPING.items():
        if prefix in test_id:
            return name
    return "Module: General Security Control"

class SecurityExcelReporter:
    def __init__(self):
        self.results = []
        self.start_time = None
        self.end_time = None

    def add_result(self, test_id, method_name, status, duration, failure_reason=""):
        module_name = extract_module_name(test_id)
        self.results.append({
            "test_id": test_id,
            "module": module_name,
            "method_name": method_name,
            "status": status,
            "duration": round(duration, 3),
            "failure_reason": failure_reason
        })

    def generate_report(self, output_path="reports/Security_Controls_Audit_Report.xlsx", base_url="http://127.0.0.1:5000"):
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb = openpyxl.Workbook()

        # Define Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)

        fill_header_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        fill_header_navy = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        fill_header_red = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        fill_header_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

        fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        font_pass = Font(name="Calibri", size=11, bold=True, color="006100")

        fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        font_fail = Font(name="Calibri", size=11, bold=True, color="9C0006")

        fill_skip = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        font_skip = Font(name="Calibri", size=11, bold=True, color="9C6500")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        total_tests = len(self.results)
        passed_tests = sum(1 for r in self.results if r["status"] == "PASSED")
        failed_tests = sum(1 for r in self.results if r["status"] == "FAILED")
        skipped_tests = sum(1 for r in self.results if r["status"] == "SKIPPED")
        pass_rate = f"{(passed_tests / total_tests * 100):.2f}%" if total_tests > 0 else "0.00%"
        total_duration = sum(r["duration"] for r in self.results)

        # -------------------------------------------------------------
        # SHEET 1: Executive Summary
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        ws_summary.views.sheetView[0].showGridLines = True

        ws_summary["A1"] = "SmileAI Enterprise Security Posture Audit Report"
        ws_summary["A1"].font = title_font

        ws_summary.cell(row=3, column=1, value="Metric").font = header_font
        ws_summary.cell(row=3, column=1).fill = fill_header_blue
        ws_summary.cell(row=3, column=2, value="Details / Value").font = header_font
        ws_summary.cell(row=3, column=2).fill = fill_header_blue

        summary_metrics = [
            ("Execution Timestamp", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Target API Base URL", base_url),
            ("Total Test Cases Evaluated", total_tests),
            ("Passed Test Cases", passed_tests),
            ("Failed Test Cases", failed_tests),
            ("Skipped Test Cases", skipped_tests),
            ("Security Posture Pass Rate", pass_rate),
            ("Total Suite Execution Time", f"{total_duration:.2f} seconds"),
            ("Audit Standard Compliance", "OWASP ASVS / ISO 27001 Baseline Passed" if failed_tests == 0 else "Action Required - Fix Failing Controls")
        ]

        for i, (m, v) in enumerate(summary_metrics, start=4):
            c1 = ws_summary.cell(row=i, column=1, value=m)
            c2 = ws_summary.cell(row=i, column=2, value=v)
            c1.font = bold_font
            c2.font = regular_font
            c1.border = thin_border
            c2.border = thin_border
            if m == "Security Posture Pass Rate":
                c2.font = font_pass if failed_tests == 0 else font_fail
                c2.fill = fill_pass if failed_tests == 0 else fill_fail

        # -------------------------------------------------------------
        # SHEET 2: Security Test Cases (All 200)
        # -------------------------------------------------------------
        ws_cases = wb.create_sheet(title="Security Test Cases")
        ws_cases.views.sheetView[0].showGridLines = True

        headers_cases = ["Test ID", "Security Module", "Test Case Method", "Status", "Duration (s)", "Remarks / Compliance Note"]
        for col_num, h_text in enumerate(headers_cases, 1):
            cell = ws_cases.cell(row=1, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = fill_header_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, r in enumerate(self.results, start=2):
            c_id = ws_cases.cell(row=row_idx, column=1, value=r["test_id"])
            c_mod = ws_cases.cell(row=row_idx, column=2, value=r["module"])
            c_meth = ws_cases.cell(row=row_idx, column=3, value=r["method_name"])
            c_stat = ws_cases.cell(row=row_idx, column=4, value=r["status"])
            c_dur = ws_cases.cell(row=row_idx, column=5, value=r["duration"])
            c_rem = ws_cases.cell(row=row_idx, column=6, value="Control Verified" if r["status"] == "PASSED" else r["failure_reason"])

            c_id.font = bold_font
            c_mod.font = regular_font
            c_meth.font = regular_font
            c_dur.font = regular_font
            c_rem.font = regular_font

            c_stat.alignment = Alignment(horizontal="center")
            if r["status"] == "PASSED":
                c_stat.fill = fill_pass
                c_stat.font = font_pass
            elif r["status"] == "FAILED":
                c_stat.fill = fill_fail
                c_stat.font = font_fail
            else:
                c_stat.fill = fill_skip
                c_stat.font = font_skip

            for col in range(1, 7):
                ws_cases.cell(row=row_idx, column=col).border = thin_border

        # -------------------------------------------------------------
        # SHEET 3: Failed Tests
        # -------------------------------------------------------------
        ws_failed = wb.create_sheet(title="Failed Tests")
        ws_failed.views.sheetView[0].showGridLines = True

        headers_failed = ["Test ID", "Security Module", "Test Case Method", "Failure Reason & Traceback"]
        for col_num, h_text in enumerate(headers_failed, 1):
            cell = ws_failed.cell(row=1, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = fill_header_red
            cell.alignment = Alignment(horizontal="center", vertical="center")

        failed_list = [r for r in self.results if r["status"] == "FAILED"]
        if not failed_list:
            ws_failed.cell(row=2, column=1, value="No test failures detected. All security controls passed!").font = font_pass
        else:
            for row_idx, r in enumerate(failed_list, start=2):
                ws_failed.cell(row=row_idx, column=1, value=r["test_id"]).font = bold_font
                ws_failed.cell(row=row_idx, column=2, value=r["module"]).font = regular_font
                ws_failed.cell(row=row_idx, column=3, value=r["method_name"]).font = regular_font
                ws_failed.cell(row=row_idx, column=4, value=r["failure_reason"]).font = font_fail

                for col in range(1, 5):
                    ws_failed.cell(row=row_idx, column=col).border = thin_border

        # -------------------------------------------------------------
        # SHEET 4: Security Module Summary
        # -------------------------------------------------------------
        ws_modules = wb.create_sheet(title="Module Summary")
        ws_modules.views.sheetView[0].showGridLines = True

        headers_modules = ["Module Name", "Total Tests", "Passed", "Failed", "Skipped", "Pass Rate (%)"]
        for col_num, h_text in enumerate(headers_modules, 1):
            cell = ws_modules.cell(row=1, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = fill_header_green
            cell.alignment = Alignment(horizontal="center", vertical="center")

        module_stats = {}
        for m_name in MODULE_MAPPING.values():
            module_stats[m_name] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}

        for r in self.results:
            m = r["module"]
            if m not in module_stats:
                module_stats[m] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}
            module_stats[m]["total"] += 1
            if r["status"] == "PASSED":
                module_stats[m]["passed"] += 1
            elif r["status"] == "FAILED":
                module_stats[m]["failed"] += 1
            else:
                module_stats[m]["skipped"] += 1

        for r_idx, (m_name, stats) in enumerate(module_stats.items(), start=2):
            tot = stats["total"]
            pas = stats["passed"]
            fai = stats["failed"]
            skp = stats["skipped"]
            prate = f"{(pas / tot * 100):.1f}%" if tot > 0 else "0.0%"

            ws_modules.cell(row=r_idx, column=1, value=m_name).font = bold_font
            ws_modules.cell(row=r_idx, column=2, value=tot).font = regular_font
            ws_modules.cell(row=r_idx, column=3, value=pas).font = regular_font
            ws_modules.cell(row=r_idx, column=4, value=fai).font = regular_font
            ws_modules.cell(row=r_idx, column=5, value=skp).font = regular_font
            
            prate_cell = ws_modules.cell(row=r_idx, column=6, value=prate)
            prate_cell.font = font_pass if fai == 0 else font_fail
            prate_cell.fill = fill_pass if fai == 0 else fill_fail

            for col in range(1, 7):
                ws_modules.cell(row=r_idx, column=col).border = thin_border

        # Adjust Column Widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        max_len = max(max_len, len(val_str.split("\n")[0]))
                sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 60)

        wb.save(output_path)
        print(f"\n[security_excel_reporter] Multi-sheet Excel report generated successfully: {output_path}")

reporter = SecurityExcelReporter()
